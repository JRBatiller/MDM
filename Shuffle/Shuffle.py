import numpy as np
from openpyxl import load_workbook
import openpyxl
import datetime


"""-----------------------------SpreadSheets---------------------------------"""
wb = load_workbook(filename = 'NDC-Data-superMOD.xlsx', data_only=True)
ws = wb["Yearly"]

dis_wb = load_workbook(filename = 'Horizontal Distance.xlsx', data_only=True)
dis_ws = dis_wb["Horizontal Distance"]


""" ----------------FUNCTIONS---------------------------------
FLOW CALCULATION
e-flow : Takes a gates flow and numbering of neighbouring empty gates and
outputs the gates effective flow

cost: Takes the effective flow matrix and the ranking matrix and outputs the
evaluated cost function of the current system

gate_eval: evaluates how many empty gates neighbour a gate. Takes a gate number
and current gate assignment and flow as input

GATE ASSIGNMENT
neigh: Evaluates which of a gates neighbours has the greatest flow. Takes a gate
number and flow assignment as input

E_insert: Inserts a '0' between 2 given gates in the position and flow arrays


"""

def e_flow(f, j):
    #f - flow
    #j - number of neighbouring gates
    if f == 0:
        return 0
    elif j == 0:
        return f
    else:
        e = f/(j+1)
        return e

def cost(E, R):
    #E - an array where each entry is the effective flow of a gate
    #R - an array where each entry is the ranking of a gate
    c = np.dot(E, R)
    return c

def gate_eval(n, G):
    #n = gate number
    # G = Array with gate assignment
    if n == 0:
        if G[n+1] == '0' or G[n+1] == '00':
            j = 1
            return j
        else:
            j = 0
            return j
    elif n == len(G) - 1:
        if G[n-1] == '0' or G[n-1] =='00':
            j = 1
            return j
        else:
            j = 0
            return j
    else:
        if (G[n+1] =='0' or G[n+1] =='00') and (G[n-1] == '0' or G[n+1] == '00'):
            j = 2
            return j
        elif G[n+1] == '0' or G[n-1] == '0' or G[n+1] =='00' or G[n-1] == '00' :
            j = 1
            return j
        else:
            j = 0
            return j

def neigh(n, F):
    # n - gate number whose largest neighbour we want to find
    # F - array whose entries are
    if n == 0:
        return n+1
    elif n == len(F) - 1:
        return n-1
    elif F[n+1] >= F[n-1]:
        return n+1
    else:
        return n-1

def E_insert(n, m, G, F):
    #n, m - gates empty gate will go between
    if n < m :
        G.insert(n+1, '00')
        F.insert(n+1, 0)
    else:
        G.insert(m+1, '00')
        F.insert(m+1, 0)


"""---------------------Initial Re-Assignment of Gates according to rank ----------
The first step is assigning gates according to rank. There are 58 departure
gates. Each region will have a corresponding rank array, where the first entry
is the closest departure for that region, the second entry the second closest,
and so on. The initial assignment will use these to assign the busiest
destinations to the closest gate.

There are 14 empty  gates
"""

"""--------------------------------RANK-------------------------------------"""
# Creating Rank Arrays
NRank = [40, 41, 39, 57, 42, 38, 56, 43, 37, 55, 44, 36, 54, 45, 35, 53, 46, 34,
52, 47, 51, 48, 50, 49]

MRank = [31, 32, 30, 33, 29, 28, 27, 26, 25, 24, 23, 22, 21]

SRank = [17, 16, 18, 15, 19, 14, 20, 0, 13, 1, 12, 2, 11, 3, 10, 4, 9, 5, 8, 6,
7]

Rank = [8, 10, 12, 14, 16, 18, 20, 21, 19, 17, 15, 13, 11, 9, 6, 4, 2, 1, 3, 5,
7, 13, 12, 11, 10, 9, 8, 7, 6, 5, 3, 1, 2, 4, 18, 15, 12, 9, 6, 3, 1, 2, 5, 8,
11, 14, 17, 20, 22, 24, 23, 21, 19, 16, 13, 10, 7, 4]
"""------------------------------TCRANK--------------------------------------"""
TSRank = []
for i in range(0, 21):
    TSRank.append(dis_ws.cell(row = 175, column = 2 + i).value)

TMRank = []
for i in range(0, 13):
    TMRank.append(dis_ws.cell(row = 179, column = 2 + i).value)

TNRank = []
for i in range(0, 24):
    TNRank.append(dis_ws.cell(row = 183, column = 2 + i).value)

TCRank = np.hstack([TSRank, TMRank, TNRank])


"""---------------------------HYPRANK---------------------------------------"""
HSRank = []
for i in range(0, 21):
    HSRank.append(dis_ws.cell(row = 192, column = 2 + i).value)

HMRank = []
for i in range(0, 13):
    HMRank.append(dis_ws.cell(row = 196, column = 2 + i).value)

HNRank = []
for i in range( 0, 24):
    HNRank.append(dis_ws.cell(row = 200, column = 2 + i).value)

HYPRank = np.hstack([HSRank, HMRank, HNRank])


"""------------------------Initial Assignment of Gates----------------------"""
#Creatin Arrays containing Gates and Flows
#Initially, each gate position is filled with zeros
Gates = []
Flows =  []
for i in range (0,58):
     Gates.append(str(0))
     Flows.append(0)
#Assigning Southern Gates
for i in range (0,16):
    Gates[SRank[i]] = ws.cell(row = 21 +i, column = 10).value
    Flows[SRank[i]] = ws.cell(row = 21 +i, column = 11).value
#Assigning Midland Gates
for i in range (0,9):
    Gates[MRank[i]] = ws.cell(row = 21 +i, column =6).value
    Flows[MRank[i]] = ws.cell(row = 21 +i, column = 7).value
#Assignging Northern Gates
for i in range (0, 13):
    Gates[NRank[i]] = ws.cell(row = 21 +i, column = 2).value
    Flows[NRank[i]] = ws.cell(row = 21 +i, column = 3).value



"""------------------Calculation of Effective Flow---------------------------
An array whose entries are the calculated effective flows for each gate need to
be made. Initially, I am choosing to only count gates directly left or right of
a gate as a neighbour Empty gates automatically get an effective flow of 0, if a
gate has 1 empty neighbiour it's flow is halved etc."""

E_flow = []
for i in range(0, len(Gates)):
    j = gate_eval(i, Gates)
    f = Flows[i]
    E_flow.append(e_flow(f, j))

#print(cost(E_flow, TCRank))


"""--------------------Assigning Empty Gates-----------------------------
Now we need to start assigning extra gates to the destinations with the greatest
moved. To keep track of gates, an an empty gate array containing current
of the empty gates is used
"""




"""-----------------------------MAGIC--------------------------------------"""

for i in range(0, 20):
    G_empty = []
    for i in range(0, len(Gates)):
        if Gates[i] == '0':
            G_empty.append(i)

    del Gates[G_empty[0]]
    del Flows[G_empty[0]]

    E_flow = []
    for i in range (0, len(Gates)):
        j = gate_eval(i, Gates)
        f = Flows[i]
        E_flow.append(e_flow(f, j))

    Ord = np.argsort(E_flow)
    N = Ord[-1]

    if gate_eval(N, Gates) == 2:
        N = Ord[-2]
        if gate_eval(N, Gates) == 2:
            N = Ord[-3]

    if N == len(Gates)-1 and gate_eval(N, Gates) ==1:
        N = Ord[-2]

    E_insert(N, neigh(N, E_flow), Gates, Flows)
    E_flow = []
    for i in range(0, len(Gates)):
        j = gate_eval(i, Gates)
        f = Flows[i]
        E_flow.append(e_flow(f, j))

print(cost(E_flow, HYPRank))

"""--------------------------Old Arrangement Cost Function-------------------"""

Gates_I = []
Flows_I = []
for i in range (0,21):
    Gates_I.append(ws.cell(row = 56+i, column = 21).value)
    Flows_I.append(ws.cell(row = 56+i, column = 22).value)
for i in range (0, 13):
    Gates_I.append(ws.cell(row = 56+i, column = 17).value)
    Flows_I.append(ws.cell(row = 56+i, column = 18).value)
for i in range (0, 24):
    Gates_I.append(ws.cell(row = 56+i, column = 13).value)
    Flows_I.append(ws.cell(row = 56+i, column = 14).value)

for i in range (0, len(Gates_I)):
    if Gates_I[i] == 0L:
        Gates_I[i] = '0'
    if Flows_I[i] == 0L:
        Flows_I[i] = 0

E_flows_I = []
for i in range (0, len(Gates_I)):
    j = gate_eval(i, Gates_I)
    f = Flows_I[i]
    E_flows_I.append(e_flow(f, j))

print(cost(E_flows_I, HYPRank))
